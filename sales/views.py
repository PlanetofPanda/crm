from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from io import BytesIO
import json

from .forms import CaptchaAuthenticationForm, CustomerForm, ImportForm, UserManagementForm
from .models import Customer
from .decorators import admin_required, sales_required


def login_view(request):
    """登录视图"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = CaptchaAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('dashboard')
    else:
        form = CaptchaAuthenticationForm()
    
    return render(request, 'login.html', {'form': form})


def logout_view(request):
    """登出视图"""
    logout(request)
    return redirect('login')


@login_required
def dashboard_view(request):
    """仪表盘视图 - 显示所有未来的任务"""
    user = request.user
    
    # 获取当前时间
    now = timezone.now()
    
    # 查询所有未来的任务
    if user.is_superuser:
        # 管理员看所有任务
        future_tasks = Customer.objects.filter(
            next_contact_time__gte=now
        ).order_by('next_contact_time').select_related('sales_rep')
    else:
        # 销售人员只看自己的任务
        future_tasks = Customer.objects.filter(
            sales_rep=user,
            next_contact_time__gte=now
        ).order_by('next_contact_time')
    
    # 按日期分组
    from itertools import groupby
    from collections import OrderedDict
    
    tasks_by_date = OrderedDict()
    for task in future_tasks:
        task_date = timezone.localtime(task.next_contact_time).date()
        if task_date not in tasks_by_date:
            tasks_by_date[task_date] = []
        tasks_by_date[task_date].append(task)
    
    # 获取今天日期用于高亮显示
    today = timezone.now().date()
    
    context = {
        'tasks_by_date': tasks_by_date,
        'today_date': today,
    }
    
    return render(request, 'dashboard.html', context)


@sales_required
def my_customers_view(request):
    """我的客户列表页"""
    user = request.user
    
    
    # 处理批量添加（所有用户）
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'batch_add':
            # 批量添加客户
            try:
                import json
                from django.utils.dateparse import parse_datetime
                
                batch_data = request.POST.get('batch_data', '[]')
                customers_data = json.loads(batch_data)
                
                created_count = 0
                skipped_count = 0
                error_messages = []
                
                for customer_data in customers_data:
                    try:
                        # 检查电话是否已存在
                        phone = customer_data.get('phone', '').strip()
                        if Customer.objects.filter(phone=phone).exists():
                            skipped_count += 1
                            continue
                        
                        # 准备客户数据
                        customer = Customer()
                        customer.name = customer_data.get('name', '').strip()
                        customer.phone = phone
                        customer.status = customer_data.get('status', 'wait_contact')
                        customer.source = customer_data.get('source', '').strip()
                        customer.city_auto = customer_data.get('city_auto', '').strip()
                        customer.region_manual = customer_data.get('region_manual', '').strip()
                        
                        # 处理创建时间
                        created_at_str = customer_data.get('created_at', '').strip()
                        if created_at_str:
                            try:
                                customer.created_at = parse_datetime(created_at_str)
                                if customer.created_at is None:
                                    # 如果解析失败，使用当前时间
                                    customer.created_at = timezone.now()
                            except:
                                customer.created_at = timezone.now()
                        else:
                            customer.created_at = timezone.now()
                        
                        # 设置销售代表（如果是普通用户批量导入，自动分配给自己）
                        if not user.is_superuser:
                            customer.sales_rep = user
                        # 管理员导入默认进入公海
                        
                        # 保存扩展数据
                        customer.extra_data = customer_data.get('extra_data', {})
                        
                        customer.save()
                        created_count += 1
                        
                    except Exception as e:
                        error_messages.append(f'导入客户 {customer_data.get("name", "未知")} 失败: {str(e)}')
                        continue
                
                # 构建提示消息
                if created_count > 0:
                    messages.success(request, f'成功导入 {created_count} 个客户')
                if skipped_count > 0:
                    messages.warning(request, f'跳过 {skipped_count} 个重复电话号码的客户')
                if error_messages:
                    for msg in error_messages[:5]:  # 最多显示5个错误
                        messages.error(request, msg)
                
                return redirect('my_customers')
                
            except json.JSONDecodeError:
                messages.error(request, '数据格式错误，请重试')
                return redirect('my_customers')
            except Exception as e:
                messages.error(request, f'批量导入失败: {str(e)}')
                return redirect('my_customers')
    
    # 处理批量操作（仅管理员）
    if request.method == 'POST' and user.is_superuser:
        action = request.POST.get('action')
        customer_ids = request.POST.getlist('customer_ids')
        
        if not customer_ids:
            messages.error(request, '请至少选择一个客户')
            return redirect('my_customers')
        
        if action == 'bulk_delete':
            # 批量删除
            count = Customer.objects.filter(id__in=customer_ids).delete()[0]
            messages.success(request, f'成功删除 {count} 个客户')
            return redirect('my_customers')
            
        elif action == 'bulk_edit':
            # 批量修改
            status = request.POST.get('status')
            sales_rep_id = request.POST.get('sales_rep')
            
            customers = Customer.objects.filter(id__in=customer_ids)
            updated_count = len(customer_ids)
            
            if status:
                customers.update(status=status)
            
            if sales_rep_id == '0':
                customers.update(sales_rep=None)
            elif sales_rep_id:
                customers.update(sales_rep_id=sales_rep_id)
            
            messages.success(request, f'成功修改 {updated_count} 个客户')
            return redirect('my_customers')
    
    # 获取筛选和排序参数
    status_filter = request.GET.get('status', '')
    city_filter = request.GET.get('city', '')
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort_by', '')
    sort_order = request.GET.get('sort_order', 'asc')
    page_number = request.GET.get('page', request.session.get('last_customer_page', 1))
    
    # 基础查询
    if user.is_superuser:
        customers = Customer.objects.all()
    else:
        customers = Customer.objects.filter(sales_rep=user)
    
    # 应用筛选
    if status_filter:
        customers = customers.filter(status=status_filter)
    if city_filter:
        customers = customers.filter(Q(city_auto__icontains=city_filter) | Q(province__icontains=city_filter))
    if search_query:
        customers = customers.filter(
            Q(name__icontains=search_query) | Q(phone__icontains=search_query)
        )
    
    # 应用排序
    if sort_by:
        # 根据排序顺序添加前缀
        order_prefix = '-' if sort_order == 'desc' else ''
        
        # 支持的排序字段
        allowed_sort_fields = {
            'name': 'name',
            'status': 'status',
            'province': 'province',
            'city_auto': 'city_auto',
            'contact_count': 'contact_count',
            'next_contact_time': 'next_contact_time',
            'last_contact_at': 'last_contact_at',
            'created_at': 'created_at',
        }
        
        if sort_by in allowed_sort_fields:
            # 对于可能为空的字段，使用nulls_last
            if sort_by in ['next_contact_time']:
                from django.db.models import F
                if sort_order == 'desc':
                    customers = customers.order_by(F(allowed_sort_fields[sort_by]).desc(nulls_last=True))
                else:
                    customers = customers.order_by(F(allowed_sort_fields[sort_by]).asc(nulls_last=True))
            else:
                customers = customers.order_by(f'{order_prefix}{allowed_sort_fields[sort_by]}')
    else:
        # 默认按创建时间倒序排列
        customers = customers.order_by('-created_at')
    
    # 分页处理
    paginator = Paginator(customers, 100)  # 每页100条记录
    customers_page = paginator.get_page(page_number)
    
    # 保存当前页码到session
    request.session['last_customer_page'] = page_number
    
    context = {
        'customers': customers_page,
        'status_choices': Customer.STATUS_CHOICES,
        'current_status': status_filter,
        'current_city': city_filter,
        'search_query': search_query,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'all_users': User.objects.filter(is_staff=True) if user.is_superuser else [],
    }
    
    return render(request, 'my_customers.html', context)


@sales_required
def customer_detail_view(request, pk=None):
    """客户详情/编辑页"""
    if pk:
        customer = get_object_or_404(Customer, pk=pk)
        # 权限检查:非管理员只能编辑自己的客户
        if not request.user.is_superuser and customer.sales_rep != request.user:
            messages.error(request, '您没有权限编辑此客户')
            return redirect('my_customers')
    else:
        customer = None
    
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            customer = form.save(commit=False)
            
            # 权限控制：普通员工只能分配给自己或公海
            if not request.user.is_superuser:
                selected_rep = form.cleaned_data.get('sales_rep')
                if selected_rep and selected_rep != request.user:
                    messages.error(request, '您只能将客户分配给自己或设为公海')
                    context = {
                        'form': form,
                        'customer': customer,
                        'is_edit': pk is not None,
                    }
                    return render(request, 'customer_detail.html', context)
            
            # 新客户自动分配给当前用户（如果没有指定销售经理）
            if not customer.pk and not customer.sales_rep and not request.user.is_superuser:
                customer.sales_rep = request.user
            
            customer.save()
            messages.success(request, '客户信息保存成功')
            # 返回到之前浏览的页码
            last_page = request.session.get('last_customer_page', 1)
            return redirect(f'/my-customers/?page={last_page}')
    else:
        form = CustomerForm(instance=customer)
    
    context = {
        'form': form,
        'customer': customer,
        'is_edit': pk is not None,
    }
    
    return render(request, 'customer_detail.html', context)


@sales_required
def high_seas_view(request):
    """公海线索页 - 显示无负责人的客户"""
    # 获取筛选参数
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    
    # 基础查询:sales_rep为空
    customers = Customer.objects.filter(sales_rep__isnull=True)
    
    # 应用筛选
    if status_filter:
        customers = customers.filter(status=status_filter)
    if search_query:
        customers = customers.filter(
            Q(name__icontains=search_query) | Q(phone__icontains=search_query)
        )
    
    # 处理认领操作
    if request.method == 'POST' and 'claim' in request.POST:
        customer_ids = request.POST.getlist('customer_ids')
        if customer_ids:
            Customer.objects.filter(id__in=customer_ids, sales_rep__isnull=True).update(
                sales_rep=request.user
            )
            messages.success(request, f'成功认领 {len(customer_ids)} 个客户')
            return redirect('high_seas')
    
    # 处理批量操作（仅管理员）
    if request.method == 'POST' and request.user.is_superuser:
        action = request.POST.get('action')
        customer_ids = request.POST.getlist('customer_ids')
        
        if not customer_ids:
            messages.error(request, '请至少选择一个客户')
            return redirect('high_seas')
        
        if action == 'bulk_delete':
            # 批量删除
            count = Customer.objects.filter(id__in=customer_ids).delete()[0]
            messages.success(request, f'成功删除 {count} 个客户')
            return redirect('high_seas')
            
        elif action == 'bulk_edit':
            # 批量修改
            status = request.POST.get('status')
            sales_rep_id = request.POST.get('sales_rep')
            
            customers_to_update = Customer.objects.filter(id__in=customer_ids)
            updated_count = len(customer_ids)
            
            if status:
                customers_to_update.update(status=status)
            
            if sales_rep_id == '0':
                # 保持公海（不修改）
                pass
            elif sales_rep_id:
                customers_to_update.update(sales_rep_id=sales_rep_id)
            
            messages.success(request, f'成功修改 {updated_count} 个客户')
            return redirect('high_seas')
    
    context = {
        'customers': customers,
        'status_choices': Customer.STATUS_CHOICES,
        'current_status': status_filter,
        'search_query': search_query,
        'all_users': User.objects.filter(is_staff=True) if request.user.is_superuser else [],
    }
    
    return render(request, 'high_seas_v2.html', context)


@sales_required
def visited_customers_view(request):
    """已到访客户页"""
    user = request.user
    
    # 筛选已到访状态的客户
    if user.is_superuser:
        customers = Customer.objects.filter(status='visited')
    else:
        customers = Customer.objects.filter(status='visited', sales_rep=user)
    
    # 搜索
    search_query = request.GET.get('search', '')
    if search_query:
        customers = customers.filter(
            Q(name__icontains=search_query) | Q(phone__icontains=search_query)
        )
    
    context = {
        'customers': customers,
        'search_query': search_query,
    }
    
    return render(request, 'visited.html', context)


@sales_required
def signed_customers_view(request):
    """已签约客户页"""
    user = request.user
    
    # 筛选已签约状态的客户
    if user.is_superuser:
        customers = Customer.objects.filter(status='signed')
    else:
        customers = Customer.objects.filter(status='signed', sales_rep=user)
    
    # 搜索
    search_query = request.GET.get('search', '')
    if search_query:
        customers = customers.filter(
            Q(name__icontains=search_query) | Q(phone__icontains=search_query)
        )
    
    context = {
        'customers': customers,
        'search_query': search_query,
    }
    
    return render(request, 'signed.html', context)


@sales_required
def key_customers_view(request):
    """重点客户列表页"""
    user = request.user
    
    # 基础查询:标记为重点客户
    if user.is_superuser:
        customers = Customer.objects.filter(is_key_customer=True)
    else:
        customers = Customer.objects.filter(
            sales_rep=user,
            is_key_customer=True
        )
    
    # 获取筛选和排序参数
    status_filter = request.GET.get('status', '')
    city_filter = request.GET.get('city', '')
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort_by', '')
    sort_order = request.GET.get('sort_order', 'asc')
    page_number = request.GET.get('page', 1)
    
    # 应用筛选
    if status_filter:
        customers = customers.filter(status=status_filter)
    if city_filter:
        customers = customers.filter(Q(city_auto__icontains=city_filter) | Q(province__icontains=city_filter))
    if search_query:
        customers = customers.filter(
            Q(name__icontains=search_query) | Q(phone__icontains=search_query)
        )
    
    # 应用排序
    if sort_by:
        order_prefix = '-' if sort_order == 'desc' else ''
        allowed_sort_fields = {
            'name': 'name',
            'status': 'status',
            'province': 'province',
            'contact_count': 'contact_count',
            'next_contact_time': 'next_contact_time',
            'created_at': 'created_at',
        }
        
        if sort_by in allowed_sort_fields:
            if sort_by in ['next_contact_time']:
                from django.db.models import F
                if sort_order == 'desc':
                    customers = customers.order_by(F(allowed_sort_fields[sort_by]).desc(nulls_last=True))
                else:
                    customers = customers.order_by(F(allowed_sort_fields[sort_by]).asc(nulls_last=True))
            else:
                customers = customers.order_by(f'{order_prefix}{allowed_sort_fields[sort_by]}')
    else:
        customers = customers.order_by('-created_at')
    
    # 分页处理
    paginator = Paginator(customers, 100)
    customers_page = paginator.get_page(page_number)
    
    context = {
        'customers': customers_page,
        'status_choices': Customer.STATUS_CHOICES,
        'current_status': status_filter,
        'current_city': city_filter,
        'search_query': search_query,
        'sort_by': sort_by,
        'sort_order': sort_order,
    }
    
    return render(request, 'key_customers.html', context)


@login_required
def get_pending_reminders_api(request):
    """获取待提醒的预约任务API"""
    user = request.user
    now = timezone.now()
    time_window = timedelta(minutes=5)
    
    # 查询5分钟内需要提醒的任务
    query = Customer.objects.filter(
        next_contact_time__gte=now - time_window,
        next_contact_time__lte=now + time_window
    )
    
    if not user.is_superuser:
        query = query.filter(sales_rep=user)
    
    reminders = []
    for customer in query:
        # 检查是否已经提醒过(使用session标记)
        session_key = f'reminded_{customer.id}_{customer.next_contact_time.strftime("%Y%m%d%H%M")}'
        if not request.session.get(session_key):
            reminders.append({
                'customer_id': customer.id,
                'customer_name': customer.name,
                'next_contact_time': customer.next_contact_time.strftime('%Y-%m-%d %H:%M'),
                'status': customer.get_status_display(),
                'phone': customer.phone,
                'notes': customer.notes or ''
            })
            # 标记已提醒(2小时后过期)
            request.session[session_key] = True
            request.session.set_expiry(7200)
    
    return JsonResponse({'reminders': reminders})


@admin_required
def settings_view(request):
    """系统设置页 - 仅管理员"""
    users = User.objects.all()
    
    # 处理用户添加
    if request.method == 'POST' and 'add_user' in request.POST:
        user_form = UserManagementForm(request.POST)
        if user_form.is_valid():
            user_form.save()
            messages.success(request, '用户添加成功')
            return redirect('settings')
    else:
        user_form = UserManagementForm()
    
    # 处理用户删除
    if request.method == 'POST' and 'delete_user' in request.POST:
        user_id = request.POST.get('user_id')
        if user_id and int(user_id) != request.user.id:
            User.objects.filter(id=user_id).delete()
            messages.success(request, '用户删除成功')
            return redirect('settings')
    
    context = {
        'users': users,
        'user_form': user_form,
    }
    
    return render(request, 'settings.html', context)


@admin_required
def export_customers_api(request):
    """导出客户数据API - 仅管理员"""
    # 获取导出类型
    export_type = request.GET.get('type', 'all')
    
    if export_type == 'signed':
        customers = Customer.objects.filter(status='signed')
        filename = '已签约客户.xlsx'
    else:
        customers = Customer.objects.all()
        filename = '全部客户.xlsx'
    
    # 创建Excel文件
    wb = Workbook()
    ws = wb.active
    ws.title = "客户数据"
    
    # 表头
    headers = ['姓名', '电话', '状态', '负责人', '线索渠道', '自动定位城市', 
               '手动填写地域', '沟通次数', '下次联系时间', '线索创建时间', '备注信息']
    ws.append(headers)
    
    # 数据行
    for customer in customers:
        row = [
            customer.name,
            customer.phone,
            customer.get_status_display(),
            customer.sales_rep.username if customer.sales_rep else '公海',
            customer.source,
            customer.city_auto,
            customer.region_manual,
            customer.contact_count,
            customer.next_contact_time.strftime('%Y-%m-%d %H:%M:%S') if customer.next_contact_time else '',
            customer.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            customer.extra_data.get('note', ''),  # 备注信息
        ]
        ws.append(row)
    
    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 返回响应
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@admin_required
def import_customers_api(request):
    """导入客户数据API - 管理员导入到公海,员工导入到私海"""
    if request.method == 'POST':
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            try:
                # 读取Excel文件
                wb = load_workbook(excel_file)
                ws = wb.active
                
                # 跳过表头
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                
                imported_count = 0
                for row in rows:
                    if not row[0] or not row[1]:  # 姓名和电话必填
                        continue
                    
                    # 创建或更新客户
                    customer, created = Customer.objects.get_or_create(
                        phone=row[1],
                        defaults={
                            'name': row[0],
                            'source': row[4] if len(row) > 4 else '',
                            'city_auto': row[5] if len(row) > 5 else '',
                            'region_manual': row[6] if len(row) > 6 else '',
                        }
                    )
                    
                    # 分配负责人
                    if request.user.is_superuser:
                        customer.sales_rep = None  # 管理员导入到公海
                    else:
                        customer.sales_rep = request.user  # 员工导入到私海
                    
                    customer.save()
                    imported_count += 1
                
                messages.success(request, f'成功导入 {imported_count} 条客户数据')
                
            except Exception as e:
                messages.error(request, f'导入失败: {str(e)}')
            
            return redirect('settings')
    
    return redirect('settings')


@admin_required
def backup_data_api(request):
    """数据备份API - 仅管理员"""
    # 导出所有客户数据
    return export_customers_api(request)
